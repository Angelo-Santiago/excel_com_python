import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

# Dados para o DataFrame de operações
dados = {
    'Mês': ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'],
    'Entrada': [10000, 12000, 15000, 13000, 11000, 14000, 16000, 15500, 14500, 12500, 17000, 18000],
    'Saída': [8000, 7000, 9000, 8500, 7500, 9500, 10000, 9800, 9200, 8300, 10500, 11000],
    'Entrada Estimada': [9500, 11500, 14000, 12500, 10500, 13500, 15500, 15000, 14000, 12000, 16500, 17500],
    'Saída Estimada': [7800, 7200, 8800, 8600, 7400, 9400, 9800, 9700, 9100, 8200, 10200, 10800],
    'Tipo de Saída': ['Dinheiro', 'Pix', 'Cartão de Crédito', 'Cartão de Débito', 'Dinheiro', 'Pix', 'Cartão de Crédito', 'Cartão de Débito', 'Dinheiro', 'Pix', 'Cartão de Crédito', 'Cartão de Débito']
}

df = pd.DataFrame(dados)
df['Lucro'] = df['Entrada'] - df['Saída']

# Dados para o DRE
dre_dados = {
    'Descrição': ['Receita Bruta', '(-) Custo das Mercadorias Vendidas (CMV)', 'Lucro Bruto', '(-) Despesas Operacionais', 'Lucro Operacional', '(-) Outras Despesas', 'Lucro Líquido'],
    'Valor': [df['Entrada'].sum(), df['Saída'].sum(), df['Entrada'].sum() - df['Saída'].sum(), 20000, 0, 0, df['Entrada'].sum() - df['Saída'].sum() - 20000]
}

df_dre = pd.DataFrame(dre_dados)

# Salvar o DataFrame de operações em um arquivo Excel
excel_path = 'dashboard_financeira.xlsx'
df.to_excel(excel_path, sheet_name='Dados Operacionais', index=False)

# Criar gráficos
plt.figure(figsize=(14, 8))

# Gráfico de barras: Entrada e Saída
plt.subplot(2, 2, 1)
plt.bar(df['Mês'], df['Entrada'], color='green', label='Entrada')
plt.bar(df['Mês'], df['Saída'], color='red', label='Saída', alpha=0.7)
plt.xlabel('Mês')
plt.ylabel('Valor')
plt.title('Entrada e Saída Mensal')
plt.xticks(rotation=45)
plt.legend()

# Gráfico de linhas: Entrada e Saída Estimada vs Real
plt.subplot(2, 2, 2)
plt.plot(df['Mês'], df['Entrada'], marker='o', label='Entrada Real')
plt.plot(df['Mês'], df['Entrada Estimada'], marker='o', linestyle='--', label='Entrada Estimada')
plt.plot(df['Mês'], df['Saída'], marker='x', label='Saída Real')
plt.plot(df['Mês'], df['Saída Estimada'], marker='x', linestyle='--', label='Saída Estimada')
plt.xlabel('Mês')
plt.ylabel('Valor')
plt.title('Entrada e Saída: Real vs Estimado')
plt.xticks(rotation=45)
plt.legend()

# Gráfico de barras: Lucro Mensal
plt.subplot(2, 2, 3)
plt.bar(df['Mês'], df['Lucro'], color='blue')
plt.xlabel('Mês')
plt.ylabel('Lucro')
plt.title('Lucro Mensal')
plt.xticks(rotation=45)

# Gráfico de setores: Tipo de Saída
tipo_saida_counts = df['Tipo de Saída'].value_counts()
plt.subplot(2, 2, 4)
plt.pie(tipo_saida_counts, labels=tipo_saida_counts.index, autopct='%1.1f%%', startangle=140)
plt.title('Distribuição dos Tipos de Saída')

# Salvar os gráficos em um buffer de memória
buf = BytesIO()
plt.tight_layout()
plt.savefig(buf, format='png')
buf.seek(0)
plt.close()

# Carregar a planilha existente
workbook = load_workbook(excel_path)

# Adicionar gráficos à planilha existente
sheet = workbook['Dados Operacionais']
image = Image(buf)
sheet.add_image(image, 'E5')

# Adicionar uma nova planilha para o DRE
dre_sheet = workbook.create_sheet(title='DRE')
for r in dataframe_to_rows(df_dre, index=False, header=True):
    dre_sheet.append(r)

# Ajustar a largura das colunas para melhor visualização
for column in dre_sheet.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    dre_sheet.column_dimensions[column_letter].width = adjusted_width

# Salvar a planilha atualizada
workbook.save(excel_path)

print("Dashboard com DRE criada com sucesso!")
